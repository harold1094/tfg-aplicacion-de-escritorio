"""Exportación a Excel."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _stringify(value: Any) -> Any:
    if hasattr(value, "value"):
        return value.value
    return value


def export_rows_to_excel(
    rows: Iterable[Mapping[str, Any]],
    file_path: str | Path,
    sheet_name: str = "Facturas",
    fieldnames: list[str] | None = None,
) -> Path:
    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    row_list = list(rows)

    if fieldnames is None:
        fieldnames = list(row_list[0].keys()) if row_list else []

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]
    worksheet.append(fieldnames)

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")

    for row in row_list:
        worksheet.append([_stringify(row.get(field, "")) for field in fieldnames])

    worksheet.freeze_panes = "A2"
    for index, field in enumerate(fieldnames, start=1):
        values = [str(row.get(field, "")) for row in row_list]
        width = max([len(field), *[len(value) for value in values]], default=10)
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 12), 40)

    workbook.save(path)
    return path

